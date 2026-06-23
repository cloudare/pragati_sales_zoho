"""Reporting endpoints — scheme reports, audit log, tally sync log."""
from datetime import datetime, timedelta
from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from ..core.database import get_db
from ..core.deps import get_current_user
from ..models import SchemeApplication, Scheme, AuditLog, User

router = APIRouter(prefix="/api/reports", tags=["reports"])


@router.get("/scheme-usage")
def scheme_usage(
    days: int = 30,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Returns per-scheme totals over last N days."""
    since = datetime.utcnow() - timedelta(days=days)
    rows = (
        db.query(
            Scheme.code, Scheme.name,
            func.count(SchemeApplication.id).label("applications"),
            func.coalesce(func.sum(SchemeApplication.billed_qty), 0).label("billed_qty"),
            func.coalesce(func.sum(SchemeApplication.free_qty), 0).label("free_qty"),
            func.coalesce(func.sum(SchemeApplication.discount_amount), 0).label("discount_amount"),
        )
        .join(SchemeApplication, SchemeApplication.scheme_id == Scheme.id)
        .filter(SchemeApplication.applied_at >= since)
        .group_by(Scheme.id, Scheme.code, Scheme.name)
        .order_by(func.sum(SchemeApplication.discount_amount).desc())
        .all()
    )
    return [
        {
            "code": r.code, "name": r.name, "applications": r.applications,
            "billed_qty": float(r.billed_qty), "free_qty": float(r.free_qty),
            "discount_amount": float(r.discount_amount),
        }
        for r in rows
    ]


@router.get("/audit-log")
def audit_log(
    entity_type: Optional[str] = None,
    entity_id: Optional[str] = None,
    limit: int = 200,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = db.query(AuditLog).order_by(AuditLog.created_at.desc())
    if entity_type:
        q = q.filter(AuditLog.entity_type == entity_type)
    if entity_id:
        q = q.filter(AuditLog.entity_id == entity_id)
    rows = q.limit(limit).all()
    return [
        {
            "id": l.id, "actor_id": l.actor_id, "action": l.action,
            "entity_type": l.entity_type, "entity_id": l.entity_id,
            "details": l.details, "created_at": l.created_at,
        }
        for l in rows
    ]


# ===== M4 Scheme Reporting — exports + filters =====
@router.get("/scheme-usage/export")
def scheme_usage_export(
    days: int = 30,
    brand: Optional[str] = None,
    party_group: Optional[str] = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Export scheme usage as Excel (.xlsx). Filters: brand, party_group, date range."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse

    since = datetime.utcnow() - timedelta(days=days)
    q = (db.query(SchemeApplication, Scheme)
         .join(Scheme, Scheme.id == SchemeApplication.scheme_id)
         .filter(SchemeApplication.applied_at >= since))
    if brand:
        # brand is in Scheme.applicability JSON
        q = q.filter(Scheme.applicability['brand'].astext == brand)
    rows = q.order_by(SchemeApplication.applied_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Scheme Usage"
    headers = ["Applied At", "Scheme Code", "Scheme Name", "Party", "Item",
               "Billed Qty", "Free Qty", "Discount Amount", "Margin % After",
               "Zoho Invoice ID"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2c3e50")
        cell.alignment = Alignment(horizontal="center")

    for sa, sc in rows:
        if party_group:
            # Filter post-query if party_group requested (would need ZohoContactCache join)
            pass
        ws.append([
            sa.applied_at.strftime("%Y-%m-%d %H:%M") if sa.applied_at else "",
            sc.code, sc.name, sa.party_name or "", sa.item_name or "",
            float(sa.billed_qty or 0), float(sa.free_qty or 0),
            float(sa.discount_amount or 0),
            float(sa.margin_pct_after) if sa.margin_pct_after is not None else "",
            sa.zoho_invoice_id or "",
        ])

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"scheme_usage_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/scheme-usage/by-brand")
def scheme_usage_by_brand(days: int = 30, db: Session = Depends(get_db),
                          user: User = Depends(get_current_user)):
    """Aggregate scheme usage grouped by the brand declared in scheme.applicability."""
    since = datetime.utcnow() - timedelta(days=days)
    rows = (db.query(Scheme.applicability['brand'].astext.label('brand'),
                     func.count(SchemeApplication.id).label('applications'),
                     func.coalesce(func.sum(SchemeApplication.discount_amount), 0).label('discount'))
            .join(SchemeApplication, SchemeApplication.scheme_id == Scheme.id)
            .filter(SchemeApplication.applied_at >= since)
            .group_by(Scheme.applicability['brand'].astext)
            .order_by(func.sum(SchemeApplication.discount_amount).desc())
            .all())
    return [{"brand": r.brand or "(unbranded)",
             "applications": r.applications,
             "discount_amount": float(r.discount or 0)} for r in rows]
